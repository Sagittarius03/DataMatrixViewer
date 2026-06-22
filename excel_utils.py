"""
Модуль для работы с Excel файлами
Сохраняет символ 0x1D (Group Separator) в Excel через замену в XML
"""
import os
import zipfile
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def save_excel_with_raw_char(file_path, data, sheet_name="DataMatrix"):
    """
    Сохраняет данные в Excel с сохранением символа 0x1D (GS)
    
    Алгоритм:
    1. Создаём Excel через openpyxl
    2. Заменяем 0x1D на маркер в данных
    3. Сохраняем как xlsx
    4. Открываем как zip и заменяем маркер на 0x1D в бинарном режиме
    
    Args:
        file_path (str): Путь для сохранения
        data (list): Список строк для записи
        sheet_name (str): Название листа
    
    Returns:
        bool: True если успешно, False если ошибка
    """
    try:
        # Создаём временный файл
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        temp_path = temp_file.name
        temp_file.close()
        
        # Создаём Excel через openpyxl
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]
        
        # # Заголовок
        # ws['A1'] = "DataMatrix Коды"
        # ws['A1'].font = Font(bold=True, size=12)
        # ws['A1'].alignment = Alignment(horizontal='center')
        # ws.merge_cells('A1:A1')
        
        # Записываем данные с заменой 0x1D на маркер
        GS_MARKER = '###GS_MARKER###'
        for i, line in enumerate(data, 1):
            temp_line = line.replace(chr(0x1D), GS_MARKER)
            ws.cell(row=i, column=1, value=temp_line)
        
        ws.column_dimensions['A'].width = 80
        wb.save(temp_path)
        wb.close()
        
        # Открываем xlsx как zip и заменяем маркер на 0x1D
        with zipfile.ZipFile(temp_path, 'r') as zip_in:
            with zipfile.ZipFile(file_path, 'w', zipfile.ZIP_DEFLATED) as zip_out:
                for item in zip_in.infolist():
                    data_in = zip_in.read(item.filename)
                    
                    # Заменяем только в XML файлах
                    if item.filename.endswith('.xml'):
                        # Бинарная замена маркера на 0x1D
                        data_in = data_in.replace(b'###GS_MARKER###', b'\x1d')
                    
                    zip_out.writestr(item, data_in)
        
        # Удаляем временный файл
        os.unlink(temp_path)
        return True
        
    except Exception as e:
        print(f"Ошибка при сохранении Excel: {e}")
        return False