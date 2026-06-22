"""
Основной модуль приложения DataMatrix CSV Viewer
Содержит главный класс, управляющий всей логикой
"""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import re
import os
import sys
from collections import defaultdict
import base64
from io import BytesIO
import threading

try:
    from PIL import Image, ImageTk
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

from logo import LOGO_BASE64, LOGO_DEVELOPER
from excel_utils import save_excel_with_raw_char
from csv_utils import read_csv_with_encoding, save_csv_with_crlf
import ui_builder
import single_instance


def get_logo_image():
    """Загружает логотип из встроенного base64"""
    if not HAS_PIL:
        return None
    
    try:
        logo_data = LOGO_BASE64.strip()
        image_data = base64.b64decode(logo_data)
        image = Image.open(BytesIO(image_data))
        image = image.resize((50, 50), Image.Resampling.LANCZOS)
        return ImageTk.PhotoImage(image)
    except Exception as e:
        print(f"Ошибка загрузки логотипа: {e}")
        return None

def get_dev_logo():
    """Загружает логотип разработчика из встроенного base64"""
    if not HAS_PIL:
        return None
    
    try:
        logo_data = LOGO_DEVELOPER.strip()
        if not logo_data:
            print("LOGO_DEVELOPER пуст")
            return None
        
        # Декодируем base64
        image_data = base64.b64decode(logo_data)
        
        # Открываем изображение
        image = Image.open(BytesIO(image_data))
        
        # Ресайзим до 20x20
        image = image.resize((20, 20), Image.Resampling.LANCZOS)
        
        # Конвертируем в PhotoImage для Tkinter
        return ImageTk.PhotoImage(image)
    except Exception as e:
        print(f"Ошибка загрузки логотипа разработчика: {e}")
        # Если не получилось, создаем простой значок-заглушку
        try:
            from PIL import ImageDraw
            img = Image.new('RGBA', (20, 20), (79, 140, 247, 255))
            draw = ImageDraw.Draw(img)
            draw.text((4, 0), "Л", fill='white')
            return ImageTk.PhotoImage(img)
        except:
            return None

class DataMatrixCSVViewer:
    """Главный класс приложения"""
    
    def __init__(self, root, initial_file=None):
        self.root = root
        self.root.title("DataMatrix CSV Viewer Pro")
        self.root.geometry("1200x850")
        self.root.configure(bg='#0a0e17')
        self.root.minsize(800, 600)
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # Данные
        self.current_file = None
        self.data = []
        self.special_rows = set()
        self.special_chars_info = {}
        self.filtered_data = []
        self.filter_mode = False
        self.is_closing = False
        
        # Загружаем логотипы
        self.logo_image = get_logo_image()
        self.dev_logo = get_dev_logo()
        
        # Настройка UI
        ui_builder.setup_styles(root)
        self.setup_ui()
        
        # Настройка горячих клавиш
        self.setup_hotkeys()
        
        # Создаем контекстное меню
        self.create_context_menu()
        
        # Запускаем сервер для единственного экземпляра
        self.start_single_instance_server()
        
        # Если передан файл - открываем его
        if initial_file and os.path.exists(initial_file):
            self.root.after(100, lambda: self.open_file(initial_file))
    
    def start_single_instance_server(self):
        """Запускает сервер для обработки команд из других экземпляров"""
        def server_thread():
            single_instance.start_server(self.handle_remote_command)
        
        thread = threading.Thread(target=server_thread, daemon=True)
        thread.start()
    
    def handle_remote_command(self, file_path):
        """Обрабатывает команду от другого экземпляра"""
        def open_file():
            if os.path.exists(file_path):
                self.open_file(file_path)
                self.root.deiconify()
                self.root.lift()
                self.root.focus_force()
        
        self.root.after(100, open_file)
    
    def on_closing(self):
        """Обработчик закрытия окна"""
        self.is_closing = True
        self.root.destroy()
    
    def open_file(self, file_path):
        """Открывает файл (используется для внешних вызовов)"""
        if not os.path.exists(file_path):
            return
        
        try:
            self.current_file = file_path
            self.update_status(f"📖 Чтение файла: {os.path.basename(file_path)}")
            self.root.update()
            
            self.data, used_encoding = read_csv_with_encoding(file_path)
            
            self.display_data()
            self.update_stats()
            self.stats_labels['file'].config(text=os.path.basename(file_path))
            self.encoding_label.config(text=f"Кодировка: {used_encoding}")
            
            special_count = len(self.special_rows)
            if special_count > 0:
                self.update_status(f"✅ Загружено {len(self.data)} записей, ⚠️ {special_count} строк с особыми символами")
            else:
                self.update_status(f"✅ Загружено {len(self.data)} записей, особых символов нет")
            
            self.root.title(f"DataMatrix CSV Viewer Pro - {os.path.basename(file_path)}")
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось открыть файл:\n{str(e)}")
            self.update_status("❌ Ошибка загрузки")
    
    def setup_hotkeys(self):
        """Настройка горячих клавиш"""
        self.root.bind('<Control-o>', lambda e: self.open_csv())
        self.root.bind('<Control-s>', lambda e: self.save_as_excel())
        self.root.bind('<Control-c>', self.copy_selected_shortcut)
        self.root.bind('<Control-f>', self.find_shortcut)
        self.root.bind('<Control-a>', self.select_all_shortcut)
        self.root.bind('<Control-b>', lambda e: self.split_by_code())
        self.root.bind('<Escape>', lambda e: self.clear_filter())
    
    def setup_ui(self):
        """Создание интерфейса"""
        main_frame = ttk.Frame(self.root, style='Primary.TFrame')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)
        
        ui_builder.create_header(main_frame, self.logo_image)
        
        commands = {
            'open': self.open_csv,
            'save_excel': self.save_as_excel,
            'split': self.split_by_code,
            'copy': self.copy_selected,
            'find': self.find_special_char,
            'clear': self.clear_all,
            'clear_filter': self.clear_filter
        }
        self.filter_var, self.filter_entry = ui_builder.create_toolbar(main_frame, commands)
        self.filter_var.trace('w', self.apply_filter)
        
        self.stats_labels = ui_builder.create_stats_bar(main_frame)
        
        self.create_main_area(main_frame)
        
        self.create_footer(main_frame)
    
    def create_main_area(self, parent):
        """Создание основной области с текстовым полем"""
        main_area = ttk.Frame(parent, style='Primary.TFrame')
        main_area.pack(fill=tk.BOTH, expand=True)
        
        text_frame = ttk.Frame(main_area, style='Secondary.TFrame')
        text_frame.pack(fill=tk.BOTH, expand=True)
        text_frame.configure(relief='solid', borderwidth=1)
        
        self.text_widget = tk.Text(
            text_frame,
            bg='#0d1524',
            fg='#c8d6e5',
            font=('Cascadia Code', 10),
            wrap=tk.NONE,
            selectbackground='#2a4a7f',
            selectforeground='#ffffff',
            insertbackground='#4f8cf7',
            relief=tk.FLAT,
            padx=15,
            pady=10,
            spacing1=1,
            spacing2=1,
            spacing3=1,
            autoseparators=False,
            maxundo=0
        )
        
        try:
            self.text_widget.tk.call('tk_getOpenFile', '-nospellcheck', '1')
        except:
            pass
        
        vsb = ttk.Scrollbar(text_frame, orient="vertical", command=self.text_widget.yview)
        hsb = ttk.Scrollbar(text_frame, orient="horizontal", command=self.text_widget.xview)
        self.text_widget.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        self.text_widget.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        
        text_frame.grid_rowconfigure(0, weight=1)
        text_frame.grid_columnconfigure(0, weight=1)
        
        self.setup_text_tags()
        
        self.text_widget.bind('<Double-1>', self.on_double_click)
    
    def setup_text_tags(self):
        """Настройка тегов для подсветки текста"""
        self.text_widget.tag_configure('normal', foreground='#c8d6e5')
        self.text_widget.tag_configure('line_num', 
            foreground='#4a5a7a',
            font=('Cascadia Code', 9)
        )
        self.text_widget.tag_configure('special_char', 
            background='#8a6d00',
            foreground='#ffdd44',
            font=('Cascadia Code', 10, 'bold')
        )
        self.text_widget.tag_configure('special_line', 
            background='#2a1f00',
            foreground='#ffdd44'
        )
        self.text_widget.tag_configure('highlight', 
            background='#1a2a4a'
        )
        self.text_widget.tag_configure('filtered', 
            foreground='#4a8f7a'
        )
        self.text_widget.tag_configure('special_marker', 
            background='#ffdd44',
            foreground='#000000',
            font=('Cascadia Code', 8, 'bold')
        )
    
    def create_context_menu(self):
        """Создание контекстного меню"""
        self.context_menu = tk.Menu(self.root, tearoff=0, bg='#141b2d', fg='#e8edf5')
        self.context_menu.add_command(label="📋 Копировать", command=self.copy_selected)
        self.context_menu.add_command(label="📋 Копировать всё", command=self.copy_all)
        self.context_menu.add_separator()
        self.context_menu.add_command(label="🔍 Найти особые символы", command=self.find_special_char)
        self.context_menu.add_command(label="🔀 Разбивка по коду", command=self.split_by_code)
        self.context_menu.add_separator()
        self.context_menu.add_command(label="📊 Сохранить в Excel", command=self.save_as_excel)
        self.context_menu.add_command(label="🧹 Очистить", command=self.clear_all)
        
        self.text_widget.bind('<Button-3>', self.show_context_menu)
    
    def show_context_menu(self, event):
        """Показывает контекстное меню"""
        self.context_menu.post(event.x_root, event.y_root)
    
    def create_footer(self, parent):
        """Создание нижней панели с логотипом разработчика"""
        footer = ttk.Frame(parent, style='Primary.TFrame')
        footer.pack(fill=tk.X, pady=(10, 0))
        
        self.status_label = ttk.Label(footer, 
            text="✅ Готов к работе | Горячие клавиши: Ctrl+O, Ctrl+S, Ctrl+C, Ctrl+F, Ctrl+B",
            style='Info.TLabel')
        self.status_label.pack(side=tk.LEFT)
        
        # Разработчики
        dev_frame = ttk.Frame(footer, style='Primary.TFrame')
        dev_frame.pack(side=tk.RIGHT)
        
        # Логотип разработчика
        if self.dev_logo:
            dev_badge = tk.Label(
                dev_frame,
                image=self.dev_logo,
                bg='#0a0e17',
                borderwidth=0
            )
            dev_badge.pack(side=tk.LEFT, padx=(0, 3))
        else:
            # Запасной вариант - эмодзи
            dev_badge = tk.Label(
                dev_frame,
                text="⚡",
                font=('Segoe UI', 12),
                bg='#0a0e17',
                fg='#4f8cf7'
            )
            dev_badge.pack(side=tk.LEFT, padx=(0, 2))
        
        # Имя разработчика
        dev_name = ttk.Label(
            dev_frame,
            text="Лев Бутаков",
            style='Info.TLabel',
            font=('Segoe UI', 8, 'italic'),
            foreground='#4f8cf7'
        )
        dev_name.pack(side=tk.LEFT)
        
        # Разделитель
        sep = ttk.Label(dev_frame, text="|", style='Info.TLabel')
        sep.pack(side=tk.LEFT, padx=5)
        
        # AI Assistant
        ai_name = ttk.Label(
            dev_frame,
            text="AI Assistant",
            style='Info.TLabel',
            font=('Segoe UI', 8, 'italic'),
            foreground='#ff6b6b'
        )
        ai_name.pack(side=tk.LEFT)
        
        # Версия
        version = ttk.Label(
            dev_frame,
            text="| v2.0",
            style='Info.TLabel',
            font=('Segoe UI', 8)
        )
        version.pack(side=tk.LEFT, padx=5)
        
        self.encoding_label = ttk.Label(footer,
            text="",
            style='Info.TLabel')
        self.encoding_label.pack(side=tk.RIGHT, padx=(10, 0))
        
    # ============ ОСТАЛЬНЫЕ МЕТОДЫ ============
    # Здесь должны быть все остальные методы из вашего файла
    # (open_csv, display_data, split_by_code, save_as_excel, copy_selected, 
    # find_special_char, on_double_click, apply_filter, clear_filter, 
    # clear_all и т.д.)
    
    def open_csv(self):
        """Открытие CSV файла"""
        file_path = filedialog.askopenfilename(
            title="Выберите CSV файл",
            filetypes=[("CSV files", "*.csv"), ("Текстовые файлы", "*.txt"), ("Все файлы", "*.*")]
        )
        
        if not file_path:
            return
        
        try:
            self.current_file = file_path
            self.update_status(f"📖 Чтение файла: {os.path.basename(file_path)}")
            self.root.update()
            
            self.data, used_encoding = read_csv_with_encoding(file_path)
            
            self.display_data()
            self.update_stats()
            self.stats_labels['file'].config(text=os.path.basename(file_path))
            self.encoding_label.config(text=f"Кодировка: {used_encoding}")
            
            special_count = len(self.special_rows)
            if special_count > 0:
                self.update_status(f"✅ Загружено {len(self.data)} записей, ⚠️ {special_count} строк с особыми символами")
            else:
                self.update_status(f"✅ Загружено {len(self.data)} записей, особых символов нет")
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось открыть файл:\n{str(e)}")
            self.update_status("❌ Ошибка загрузки")
    
    def display_data(self):
        """Отображение данных с подсветкой особых символов"""
        self.text_widget.delete('1.0', tk.END)
        
        if not self.data:
            return
        
        self.special_rows = set()
        self.special_chars_info = {}
        display_data = self.filtered_data if self.filter_mode else self.data
        
        for idx, line in enumerate(display_data, 1):
            line_num = f"{idx:4d} "
            self.text_widget.insert(tk.END, line_num, 'line_num')
            
            special_chars = self.find_special_chars_in_line(line)
            
            if special_chars:
                if self.filter_mode:
                    orig_idx = self.data.index(line) + 1 if line in self.data else idx
                else:
                    orig_idx = idx
                self.special_rows.add(orig_idx)
                self.special_chars_info[orig_idx] = special_chars
                
                self.insert_line_with_special_highlight(line, special_chars)
            else:
                if self.filter_mode:
                    self.text_widget.insert(tk.END, line + '\n', 'filtered')
                else:
                    self.text_widget.insert(tk.END, line + '\n', 'normal')
        
        self.text_widget.see('1.0')
    
    def find_special_chars_in_line(self, line):
        """Находит особые символы в строке"""
        special_chars = []
        for i, char in enumerate(line):
            code = ord(char)
            if code < 32 and code not in [9, 10, 13]:
                special_chars.append((i, char, code, 'control'))
            elif code == 127:
                special_chars.append((i, char, code, 'del'))
            elif code > 127 and not char.isprintable():
                special_chars.append((i, char, code, 'unicode'))
        return special_chars
    
    def insert_line_with_special_highlight(self, line, special_chars):
        """Вставляет строку с подсветкой особых символов"""
        special_positions = {pos: (char, code, type) for pos, char, code, type in special_chars}
        
        pos = 0
        for special_pos in sorted(special_positions.keys()):
            if special_pos > pos:
                self.text_widget.insert(tk.END, line[pos:special_pos], 'special_line')
            
            char, code, char_type = special_positions[special_pos]
            
            if char_type == 'control':
                display_char = '␀'
            elif char_type == 'del':
                display_char = '␡'
            else:
                display_char = '�'
            
            self.text_widget.insert(tk.END, display_char, 'special_marker')
            pos = special_pos + 1
        
        if pos < len(line):
            self.text_widget.insert(tk.END, line[pos:], 'special_line')
        
        self.text_widget.insert(tk.END, '\n')
    
    def update_stats(self):
        """Обновление статистики"""
        total = len(self.data)
        special = len(self.special_rows)
        
        self.stats_labels['total'].config(text=str(total))
        self.stats_labels['special'].config(text=str(special))
        
        if special > 0:
            self.stats_labels['special'].config(foreground='#ffdd44')
        else:
            self.stats_labels['special'].config(foreground='#4f8cf7')
    
    def update_status(self, message):
        """Обновление статуса"""
        self.status_label.config(text=message)
        self.root.update_idletasks()
    
    def split_by_code(self):
        """Разбивка файла по кодам Data Matrix"""
        if not self.data:
            messagebox.showwarning("Предупреждение", "Сначала загрузите CSV файл")
            return
        
        format_choice = messagebox.askquestion(
            "Выбор формата",
            "Сохранять в формате Excel (.xlsx)?\n\n"
            "Нажмите 'Да' для Excel, 'Нет' для CSV",
            icon='question'
        )
        use_excel = format_choice == 'yes'
        
        output_dir = filedialog.askdirectory(
            title="Выберите папку для сохранения разбитых файлов"
        )
        
        if not output_dir:
            return
        
        try:
            self.update_status("🔀 Выполняется разбивка по кодам...")
            self.root.update()
            
            pattern = re.compile(r"010468101320(\d{7})")
            
            groups = defaultdict(list)
            errors = 0
            
            for line in self.data:
                match = pattern.search(line)
                if match:
                    code = match.group(1)
                    groups[code].append(line)
                else:
                    errors += 1
            
            split_dir = os.path.join(output_dir, "split_files")
            os.makedirs(split_dir, exist_ok=True)
            
            saved_files = []
            
            if use_excel:
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import Font
                    import tempfile
                    import zipfile
                    
                    for code, lines in groups.items():
                        file_path = os.path.join(split_dir, f"codes_{code}.xlsx")
                        
                        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
                        temp_path = temp_file.name
                        temp_file.close()
                        
                        wb = Workbook()
                        ws = wb.active
                        ws.title = f"Code_{code}"[:31]
                        
                        ws['A1'] = f"DataMatrix Коды - {code}"
                        ws['A1'].font = Font(bold=True, size=12)
                        
                        GS_MARKER = '###GS_MARKER###'
                        for i, line in enumerate(lines, 2):
                            temp_line = line.replace(chr(0x1D), GS_MARKER)
                            ws.cell(row=i, column=1, value=temp_line)
                        
                        ws.column_dimensions['A'].width = 80
                        wb.save(temp_path)
                        wb.close()
                        
                        with zipfile.ZipFile(temp_path, 'r') as zip_in:
                            with zipfile.ZipFile(file_path, 'w', zipfile.ZIP_DEFLATED) as zip_out:
                                for item in zip_in.infolist():
                                    data_in = zip_in.read(item.filename)
                                    if item.filename.endswith('.xml'):
                                        data_in = data_in.replace(b'###GS_MARKER###', b'\x1d')
                                    zip_out.writestr(item, data_in)
                        
                        os.unlink(temp_path)
                        saved_files.append((code, len(lines), file_path))
                        
                except Exception as e:
                    messagebox.showwarning(
                        "Ошибка Excel",
                        f"Ошибка при сохранении в Excel:\n{str(e)}\n\n"
                        "Будет использован формат CSV."
                    )
                    use_excel = False
            
            if not use_excel:
                for code, lines in groups.items():
                    file_path = os.path.join(split_dir, f"codes_{code}.csv")
                    
                    with open(file_path, 'w', encoding='utf-8-sig', newline='') as f:
                        for line in lines:
                            cleaned = []
                            for char in line:
                                code_ord = ord(char)
                                if code_ord == 0x1D or (32 <= code_ord <= 126) or code_ord in [9, 10, 13]:
                                    cleaned.append(char)
                                elif code_ord >= 128:
                                    cleaned.append(char)
                                else:
                                    cleaned.append(' ')
                            
                            clean_line = ''.join(cleaned)
                            
                            if '"' in clean_line or ',' in clean_line or '\n' in clean_line:
                                clean_line = '"' + clean_line.replace('"', '""') + '"'
                            
                            f.write(clean_line + '\r\n')
                    
                    saved_files.append((code, len(lines), file_path))
            
            result_text = self.build_result_message(groups, errors, saved_files, use_excel)
            messagebox.showinfo("Разбивка завершена", result_text)
            
            self.update_status(f"✅ Разбивка завершена: {len(groups)} кодов, сохранено в {split_dir}")
            
            if messagebox.askyesno("Открыть папку", "Открыть папку с результатами?"):
                os.startfile(split_dir)
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось выполнить разбивку:\n{str(e)}")
            self.update_status("❌ Ошибка разбивки")
    
    def build_result_message(self, groups, errors, saved_files, use_excel):
        """Формирует сообщение о результате разбивки"""
        result_text = f"✅ Разбивка завершена!\n\n"
        result_text += f"📊 Всего кодов: {len(groups)}\n"
        result_text += f"📄 Всего строк: {sum(len(lines) for lines in groups.values())}\n"
        result_text += f"📁 Формат: {'Excel (.xlsx) с сохранением 0x1D' if use_excel else 'CSV (.csv)'}\n"
        
        if errors > 0:
            result_text += f"⚠️ Строк без кода: {errors}\n\n"
        else:
            result_text += f"✅ Все строки содержат коды\n\n"
        
        result_text += "📋 Созданные файлы:\n"
        for code, count, file_path in saved_files[:20]:
            ext = 'xlsx' if use_excel else 'csv'
            result_text += f"  • codes_{code}.{ext} — {count} строк\n"
        
        if len(saved_files) > 20:
            result_text += f"  ... и еще {len(saved_files) - 20} файлов"
        
        return result_text
    
    def save_as_excel(self):
        """Сохранение в Excel с сохранением 0x1D"""
        if not self.data:
            messagebox.showwarning("Предупреждение", "Сначала загрузите CSV файл")
            return
        
        file_path = filedialog.asksaveasfilename(
            title="Сохранить как Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("Все файлы", "*.*")]
        )
        
        if not file_path:
            return
        
        try:
            self.update_status("💾 Создание Excel файла...")
            self.root.update()
            
            if save_excel_with_raw_char(file_path, self.data):
                self.update_status(f"✅ Сохранено: {os.path.basename(file_path)}")
                
                if messagebox.askyesno("Успех", 
                    f"Файл сохранен:\n{file_path}\n\nОткрыть файл в Excel?"):
                    os.startfile(file_path)
            else:
                messagebox.showwarning(
                    "Предупреждение",
                    "Не удалось сохранить в Excel с сохранением 0x1D.\n"
                    "Файл будет сохранен в CSV формате."
                )
                self.save_as_csv()
                
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить файл:\n{str(e)}")
            self.update_status("❌ Ошибка сохранения")
    
    def save_as_csv(self):
        """Сохранение в CSV с Windows окончаниями строк"""
        if not self.data:
            return
        
        file_path = filedialog.asksaveasfilename(
            title="Сохранить как CSV",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("Все файлы", "*.*")]
        )
        
        if not file_path:
            return
        
        try:
            if save_csv_with_crlf(file_path, self.data):
                self.update_status(f"✅ Сохранено: {os.path.basename(file_path)}")
                
                if messagebox.askyesno("Успех", 
                    f"Файл сохранен:\n{file_path}\n\nОткрыть файл?"):
                    os.startfile(file_path)
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить файл:\n{str(e)}")
    
    def copy_selected(self):
        """Копирование выделенного текста"""
        try:
            selected = self.text_widget.get(tk.SEL_FIRST, tk.SEL_LAST)
            if selected:
                self.root.clipboard_clear()
                self.root.clipboard_append(selected)
                lines = len(selected.split(chr(10)))
                self.update_status(f"📋 Скопировано {lines} строк")
        except tk.TclError:
            try:
                cursor_pos = self.text_widget.index(tk.INSERT)
                line_start = f"{cursor_pos.split('.')[0]}.0"
                line_end = f"{cursor_pos.split('.')[0]}.end"
                line_text = self.text_widget.get(line_start, line_end).strip()
                if line_text:
                    self.root.clipboard_clear()
                    self.root.clipboard_append(line_text)
                    self.update_status("📋 Скопирована текущая строка")
            except:
                messagebox.showinfo("Информация", "Выделите текст для копирования")
    
    def copy_selected_shortcut(self, event):
        """Обработчик Ctrl+C"""
        self.copy_selected()
        return "break"
    
    def select_all_shortcut(self, event):
        """Обработчик Ctrl+A"""
        self.text_widget.tag_add('sel', '1.0', 'end')
        return "break"
    
    def copy_all(self):
        """Копирование всего текста"""
        text = self.text_widget.get('1.0', tk.END).strip()
        if text:
            self.root.clipboard_clear()
            self.root.clipboard_append(text)
            self.update_status(f"📋 Скопировано {len(text.split(chr(10)))} строк")
    
    def find_special_char(self):
        """Поиск особых символов"""
        if not self.data:
            messagebox.showinfo("Информация", "Сначала загрузите файл")
            return
        
        if not self.special_rows:
            messagebox.showinfo("✅ Отлично!", "Особые символы не обнаружены")
            return
        
        dialog = tk.Toplevel(self.root)
        dialog.title("🔍 Найденные особые символы")
        dialog.geometry("800x600")
        dialog.configure(bg='#0a0e17')
        dialog.transient(self.root)
        dialog.grab_set()
        
        header = ttk.Frame(dialog, style='Secondary.TFrame')
        header.pack(fill=tk.X, padx=15, pady=15)
        
        ttk.Label(header, 
            text=f"🔍 Найдено {len(self.special_rows)} строк с особыми символами",
            style='Title.TLabel').pack()
        
        text_frame = ttk.Frame(dialog, style='Secondary.TFrame')
        text_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 15))
        text_frame.configure(relief='solid', borderwidth=1)
        
        display_text = tk.Text(text_frame,
            bg='#0d1524',
            fg='#c8d6e5',
            font=('Cascadia Code', 9),
            wrap=tk.NONE,
            selectbackground='#2a4a7f',
            selectforeground='white',
            relief=tk.FLAT,
            padx=10,
            pady=10
        )
        
        vsb = ttk.Scrollbar(text_frame, orient="vertical", command=display_text.yview)
        hsb = ttk.Scrollbar(text_frame, orient="horizontal", command=display_text.xview)
        display_text.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        display_text.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        
        text_frame.grid_rowconfigure(0, weight=1)
        text_frame.grid_columnconfigure(0, weight=1)
        
        display_text.tag_configure('special', background='#2a1f00', foreground='#ffdd44')
        display_text.tag_configure('marker', background='#ffdd44', foreground='#000000', font=('Cascadia Code', 9, 'bold'))
        
        for idx in sorted(self.special_rows):
            if idx <= len(self.data):
                line = self.data[idx - 1]
                display_text.insert(tk.END, f"{idx:4d}: ", '')
                
                special_chars = self.find_special_chars_in_line(line)
                if special_chars:
                    pos = 0
                    for special_pos, char, code, char_type in special_chars:
                        if special_pos > pos:
                            display_text.insert(tk.END, line[pos:special_pos], '')
                        
                        if char_type == 'control':
                            marker = '␀'
                        elif char_type == 'del':
                            marker = '␡'
                        else:
                            marker = '�'
                        
                        display_text.insert(tk.END, marker, 'marker')
                        pos = special_pos + 1
                    
                    if pos < len(line):
                        display_text.insert(tk.END, line[pos:], '')
                else:
                    display_text.insert(tk.END, line, '')
                
                display_text.insert(tk.END, '\n')
        
        display_text.configure(state='disabled')
        
        btn_frame = ttk.Frame(dialog, style='Primary.TFrame')
        btn_frame.pack(fill=tk.X, padx=15, pady=(0, 15))
        
        ttk.Button(btn_frame, text="✕ Закрыть", command=dialog.destroy,
                  style='Accent.TButton').pack(side=tk.RIGHT, padx=3)
    
    def on_double_click(self, event):
        """Обработчик двойного клика"""
        try:
            index = self.text_widget.index(tk.CURRENT)
            line_num = int(index.split('.')[0])
            
            if line_num <= len(self.data):
                line = self.data[line_num - 1]
                special_chars = self.find_special_chars_in_line(line)
                
                if special_chars:
                    detail = f"📄 Строка {line_num}\n"
                    detail += f"📏 Длина: {len(line)}\n"
                    detail += f"⚠️ Особые символы ({len(special_chars)}):\n\n"
                    
                    for i, char, code, char_type in special_chars:
                        char_name = {
                            'control': 'Управляющий',
                            'del': 'DELETE',
                            'unicode': 'Unicode непечатаемый'
                        }.get(char_type, 'Неизвестный')
                        
                        detail += f"  • Позиция {i}: {char_name} (код: {code}, HEX: {hex(code)})\n"
                        detail += f"    Символ: {repr(char)}\n"
                    
                    messagebox.showinfo("Информация о строке", detail)
        except:
            pass
    
    def apply_filter(self, *args):
        """Применение фильтра к данным"""
        filter_text = self.filter_var.get().strip()
        
        if not filter_text:
            self.clear_filter()
            return
        
        self.filter_mode = True
        self.filtered_data = [
            line for line in self.data 
            if filter_text.lower() in line.lower()
        ]
        
        if self.filtered_data:
            self.display_data()
            self.update_status(f"🔍 Найдено {len(self.filtered_data)} строк (фильтр: '{filter_text}')")
        else:
            self.text_widget.delete('1.0', tk.END)
            self.text_widget.insert(tk.END, f"🔍 Ничего не найдено по запросу: '{filter_text}'")
            self.update_status(f"🔍 Ничего не найдено: '{filter_text}'")
    
    def clear_filter(self):
        """Очистка фильтра"""
        self.filter_var.set("")
        self.filter_mode = False
        self.filtered_data = []
        self.display_data()
        self.update_stats()
        self.update_status("✅ Фильтр сброшен")
    
    def find_shortcut(self, event):
        """Обработчик Ctrl+F"""
        self.filter_entry.focus()
        return "break"
    
    def clear_all(self):
        """Очистка всех данных"""
        if self.data and not messagebox.askyesno("Подтверждение", "Очистить все данные?"):
            return
        
        self.text_widget.delete('1.0', tk.END)
        self.data = []
        self.special_rows = set()
        self.special_chars_info = {}
        self.filtered_data = []
        self.filter_mode = False
        self.filter_var.set("")
        self.current_file = None
        
        self.stats_labels['total'].config(text="0")
        self.stats_labels['special'].config(text="0")
        self.stats_labels['file'].config(text="Не загружен")
        self.encoding_label.config(text="")
        
        self.update_status("✅ Данные очищены")